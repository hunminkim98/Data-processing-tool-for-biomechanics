import os
import glob
import tkinter as tk
from tkinter import filedialog, simpledialog, ttk
from openpyxl import load_workbook

def process_files(folder_path, header_rows, delete_rows_dict):
    xlsx_files = glob.glob(os.path.join(folder_path, '**', '*.xlsx'), recursive=True)
    xlsx_files = sorted(xlsx_files)

    # Create "Section" folder if it doesn't exist
    section_folder = os.path.join(folder_path, "Section")
    os.makedirs(section_folder, exist_ok=True)

    for file in xlsx_files:
        file_name = os.path.basename(file)
        delete_rows = delete_rows_dict[file_name]

        wb = load_workbook(file)

        for sheet in wb.sheetnames:
            ws = wb[sheet]

            # Delete data rows
            for _ in range(delete_rows):
                ws.delete_rows(header_rows + 1)

        # Save the processed file in the "Section" folder
        new_file_name = os.path.splitext(file_name)[0] + "_Section.xlsx"
        new_file_path = os.path.join(section_folder, new_file_name)
        wb.save(new_file_path)

def on_submit():
    folder_path = entry_folder.get()
    header_rows = int(entry_header.get())

    xlsx_files = glob.glob(os.path.join(folder_path, '**', '*.xlsx'), recursive=True)
    xlsx_files = sorted(xlsx_files)

    # Create a new window to display the file list and input fields
    input_window = tk.Toplevel(window)
    input_window.title("Delete Rows")
    input_window.geometry("600x400")  # 창 크기 설정

    # Create a frame for the canvas and scrollbar
    frame_canvas = ttk.Frame(input_window)
    frame_canvas.pack(fill="both", expand=True)

    # Create a canvas and scrollbar for the file list
    canvas = tk.Canvas(frame_canvas)
    scrollbar = ttk.Scrollbar(frame_canvas, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")
        )
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    delete_rows_dict = {}
    for file in xlsx_files:
        file_name = os.path.basename(file)  # 파일 이름 추출
        label_file = tk.Label(scrollable_frame, text=file_name)  # 파일 이름으로 레이블 생성
        label_file.pack()

        entry_delete_rows = tk.Entry(scrollable_frame)
        entry_delete_rows.pack()

        delete_rows_dict[file_name] = entry_delete_rows  # 파일 이름을 키로 사용

    def on_process():
        for file_name, entry in delete_rows_dict.items():
            delete_rows = int(entry.get())
            delete_rows_dict[file_name] = delete_rows

        process_files(folder_path, header_rows, delete_rows_dict)
        input_window.destroy()
        label_status['text'] = 'Processing completed!'

    button_process = tk.Button(input_window, text="Process Files", command=on_process)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    button_process.pack()

def browse_folder():
    folder_path = filedialog.askdirectory()
    entry_folder.delete(0, tk.END)
    entry_folder.insert(tk.END, folder_path)

# Create the main window
window = tk.Tk()
window.title("Excel Data Processor")

# Folder path input
label_folder = tk.Label(window, text="Folder Path:")
label_folder.pack()
entry_folder = tk.Entry(window, width=50)
entry_folder.pack()
button_browse = tk.Button(window, text="Browse", command=browse_folder)
button_browse.pack()

# Header rows input
label_header = tk.Label(window, text="Number of Header Rows:")
label_header.pack()
entry_header = tk.Entry(window)
entry_header.pack()

# Submit button
button_submit = tk.Button(window, text="Select Files", command=on_submit)
button_submit.pack()

# Status label
label_status = tk.Label(window, text="")
label_status.pack()

window.mainloop()