"""
This script written to serve sections data.
"""
import os
import openpyxl
from tkinter import Tk
from tkinter.filedialog import askdirectory
import numpy as np
import matplotlib.pyplot as plt

def process_file_starting(file_path, start_row, start_col, header_rows, header_cols):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    initial_z = sheet.cell(row=start_row, column=start_col).value
    start_section_row = None
    
    for row in range(start_row, sheet.max_row + 1):
        z = sheet.cell(row=row, column=start_col).value
        if z is not None and initial_z is not None and z - initial_z >= 0.02:
            start_section_row = row
            break
    
    if start_section_row is None:
        print(f"파일 {file_path}에서 0.02가 증가한 지점을 찾을 수 없습니다.")
        return
    
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    
    for row in range(1, header_rows + 1):
        for col in range(1, sheet.max_column + 1):
            output_sheet.cell(row=row, column=col).value = sheet.cell(row=row, column=col).value
    
    output_row = header_rows + 1
    for row in range(start_section_row, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            output_sheet.cell(row=output_row, column=col).value = sheet.cell(row=row, column=col).value
        output_row += 1
    
    output_folder = os.path.join(os.path.dirname(file_path), "section")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    output_file_name = os.path.splitext(os.path.basename(file_path))[0] + "_section.xlsx"
    output_file_path = os.path.join(output_folder, output_file_name)
    output_wb.save(output_file_path)

def process_file_section(file_path, start_row, start_col, header_rows, header_cols, plot_endpoints):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    data = []
    for col in range(start_col, sheet.max_column + 1):
        column_data = [sheet.cell(row=row, column=col).value for row in range(start_row, sheet.max_row + 1)]
        data.append(column_data)
    
    endpoints = find_endpoints(data, plot_endpoints)
    max_endpoint = max(endpoints)
    
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    
    for row in range(1, header_rows + 1):
        for col in range(1, sheet.max_column + 1):
            output_sheet.cell(row=row, column=col).value = sheet.cell(row=row, column=col).value
    
    for col in range(1, sheet.max_column + 1):
        for row in range(start_row, start_row + max_endpoint):
            output_sheet.cell(row=row-start_row+header_rows+1, column=col).value = sheet.cell(row=row, column=col).value
    
    output_directory = os.path.dirname(file_path)
    output_folder = os.path.join(output_directory, "Phase")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    output_file_name = os.path.splitext(os.path.basename(file_path))[0] + "_Phase.xlsx"
    output_file_path = os.path.join(output_folder, output_file_name)
    output_wb.save(output_file_path)

def find_endpoints(data, plot_endpoints):
    selected_data = np.array(data)
    endpoints = []
    
    for i in range(selected_data.shape[0]):
        column_data = selected_data[i]
        endpoint = 0
        
        for j in range(1, len(column_data)):
            if column_data[j] <= column_data[j-1] or (column_data[j] > column_data[j-1] and column_data[j] - column_data[j-1] < 0.0005):
                endpoint = j
                break
        
        endpoints.append(endpoint)


    if plot_endpoints == True:
        # 기울기 그래프 그리기
        plt.figure(figsize=(10, 6))
        for i in range(selected_data.shape[0]):
            plt.plot(selected_data[i], label=f'Column {i+1}')
            if endpoints[i] < len(selected_data[i]):
                plt.axvline(x=endpoints[i], color='r', linestyle='--', label=f'Endpoint {i+1}')
        plt.xlabel('Index')
        plt.ylabel('Value')
        plt.title('Data Peaks')
        plt.legend()
        plt.tight_layout()
        plt.show()
    else:
        pass

    return endpoints
def process_excel_files():
    root = Tk()
    root.withdraw()
    directory = askdirectory(title="Select Directory")
    
    start_row = int(input("시작 행 번호를 입력하세요: "))
    start_col = int(input("시작 열 번호를 입력하세요: "))
    
    # 입력받은 행과 열에 해당하는 셀의 값을 출력
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(directory, filename)
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            cell_value = sheet.cell(row=start_row, column=start_col).value
            print(f"{filename} - 선택한 셀 ({start_row}, {start_col})의 값: {cell_value}")
    
    confirm = input("선택한 셀의 값이 맞습니까? (y/n): ")
    if confirm.lower() != 'y':
        print("프로그램을 종료합니다.")
        return
    
    header_rows = int(input("헤더 행 개수를 입력하세요: "))
    header_cols = int(input("헤더 열 개수를 입력하세요: "))
    
    find_starting_point = input("시작 지점을 찾으시겠습니까? (y/n): ")
    if find_starting_point.lower() == 'y':
        for filename in os.listdir(directory):
            if filename.endswith(".xlsx"):
                file_path = os.path.join(directory, filename)
                process_file_starting(file_path, start_row, start_col, header_rows, header_cols)
    else:
        divide_sections = input("구간을 나누시겠습니까? (y/n): ")
        if divide_sections.lower() == 'y':
            plot_endpoints = input("구간 그래프를 출력하시겠습니까? (y/n): ")
            if plot_endpoints.lower() == 'y':
                plot_endpoints = True
            for filename in os.listdir(directory):
                if filename.endswith(".xlsx"):
                    file_path = os.path.join(directory, filename)
                    process_file_section(file_path, start_row, start_col, header_rows, header_cols, plot_endpoints=plot_endpoints)
        else:
            print("프로그램을 종료합니다.")



# 사용 예시
process_excel_files()